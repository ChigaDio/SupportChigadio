# -*- coding: utf-8 -*-
"""
pythonSrc/spreadsheet_io.py

Enum / ClassDataID を Excel(.xlsx) / CSV で下ごしらえ・一括投入するための
インポート・エクスポート機能。データ量が増えてくると1行ずつのUI入力が
ボトルネックになりやすいため、Excelでまとめて作成→一括インポートできる
ようにする。

対象と1行の意味:
- enum          : 1行 = 1つの列挙値 {property, value, description}
- class_data_id : 1行 = 1つのデータレコード（columns定義に基づく動的な列。
                  先頭列は id / enum_property 固定）

設計方針（重要）:
実際の保存処理（enumの`sync_prefill_dependents`カスケードや、
class_data_idの`columns`正規化・prefill同期など）は、それぞれの
既存エンドポイント（/api/enum/<name> POST, /api/class-data-id/<name> POST）
にしか実装されていない。本モジュールで直接JSONファイルへ書き込んで
しまうと、これらの重要な副作用が抜け落ちてデータ不整合の原因になる。
そのため、アップロードされたスプレッドシートをパースしたら、
`app.test_client()` を使って既存の保存エンドポイントをそのまま
プロセス内で呼び出す方式にした（pythonSrc/generate_all.py と同じ考え方）。
これにより、本モジュールは「スプレッドシート ⇔ JSON配列の変換」だけに
専念でき、保存ロジック自体の二重実装・drift を避けられる。

.xlsx の読み書きは openpyxl に依存する。未インストールの環境でも
CSVインポート・エクスポートだけは動作するよう、openpyxl 不在時は
.xlsx関連の呼び出し時にのみエラーを返すようにしている
（アプリ全体の起動を妨げない）。
"""
import csv
import io
import json
import os

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

APP = None
DATA_DIR = None

_NUMERIC_TYPES = {"int", "uint", "short", "long", "byte"}
_FLOAT_TYPES = {"float", "double", "decimal"}
_VECTOR_TYPES = {"vector2", "vector3"}


def init(app, data_dir):
    global APP, DATA_DIR
    APP = app
    DATA_DIR = data_dir


# ---------------------------------------------------------------
# 共通: dictのリスト ⇔ スプレッドシート
# ---------------------------------------------------------------

def _cell_to_str(value):
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value


def rows_to_csv_bytes(headers, rows):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for row in rows:
        writer.writerow([_cell_to_str(row.get(h)) for h in headers])
    # ExcelでBOM無しUTF-8 CSVを開くと文字化けするため、BOM付きで出力する
    return buf.getvalue().encode("utf-8-sig")


def rows_to_xlsx_bytes(headers, rows, sheet_title="Sheet1"):
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxlがインストールされていないため.xlsx出力はできません（pip install openpyxl）")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (sheet_title or "Sheet1")[:31]
    ws.append(list(headers))
    for row in rows:
        ws.append([_cell_to_str(row.get(h)) for h in headers])
    for idx, h in enumerate(headers, start=1):
        candidates = [len(str(h))] + [len(str(_cell_to_str(row.get(h)))) for row in rows]
        width = min(max(max(candidates) + 2, 8), 60) if candidates else 8
        ws.column_dimensions[get_column_letter(idx)].width = width
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def read_uploaded_table(file_storage):
    """アップロードされたファイル(werkzeug FileStorage)を読み、
    (headers, [dict, ...]) を返す。拡張子で .csv / .xlsx を判定する。"""
    filename = (file_storage.filename or "").lower()
    if filename.endswith(".xlsx"):
        if not _HAS_OPENPYXL:
            raise RuntimeError("openpyxlがインストールされていないため.xlsxは読み込めません（pip install openpyxl）")
        wb = openpyxl.load_workbook(file_storage, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        first = next(rows_iter, None)
        headers = [str(h) if h is not None else "" for h in (first or [])]
        records = []
        for raw in rows_iter:
            if raw is None or all(v is None for v in raw):
                continue
            records.append({headers[i]: raw[i] for i in range(len(headers)) if i < len(raw)})
        return headers, records

    content = file_storage.read()
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("shift_jis", errors="replace")
    reader = csv.reader(io.StringIO(text))
    all_rows = list(reader)
    if not all_rows:
        return [], []
    headers = all_rows[0]
    records = []
    for raw in all_rows[1:]:
        if not any((c or "").strip() for c in raw):
            continue
        records.append({headers[i]: raw[i] for i in range(len(headers)) if i < len(raw)})
    return headers, records


def _coerce_scalar(value, type_name):
    if value is None:
        return None
    sval = str(value).strip()
    if type_name in _NUMERIC_TYPES:
        if sval == "":
            return 0
        try:
            return int(float(sval))
        except ValueError:
            return 0
    if type_name in _FLOAT_TYPES:
        if sval == "":
            return 0.0
        try:
            return float(sval)
        except ValueError:
            return 0.0
    if type_name == "bool":
        if isinstance(value, bool):
            return value
        return sval.lower() in ("true", "1", "yes", "on")
    if type_name in _VECTOR_TYPES:
        if sval == "":
            return []
        try:
            if sval.startswith("["):
                return json.loads(sval)
            return [float(x) for x in sval.split(",")]
        except (ValueError, json.JSONDecodeError):
            return []
    # enum名/class名/classID名/string等はそのまま文字列として扱う
    return sval


def coerce_cell(value, type_name):
    """スプレッドシートのセル値を、列定義のtype（配列は末尾'[]'）に応じて
    JSON互換の値へ変換する。既知の複合型以外はベストエフォートで文字列
    またはJSONとして解釈する。"""
    if value is None or value == "":
        return None
    if not isinstance(type_name, str):
        return _cell_to_str(value)
    if type_name.endswith("[]"):
        base = type_name[:-2]
        sval = str(value).strip()
        if sval.startswith("["):
            try:
                arr = json.loads(sval)
            except json.JSONDecodeError:
                arr = []
        else:
            arr = [x for x in sval.split("|") if x.strip() != ""]
        return [_coerce_scalar(v, base) for v in arr]
    return _coerce_scalar(value, type_name)


# ---------------------------------------------------------------
# Enum
# ---------------------------------------------------------------

_ENUM_HEADERS = ["property", "value", "description"]


def export_enum(name, fmt):
    client = APP.test_client()
    res = client.get(f"/api/enum/{name}")
    if res.status_code != 200:
        return None
    values = res.get_json() or []
    if fmt == "xlsx":
        return rows_to_xlsx_bytes(_ENUM_HEADERS, values, name)
    return rows_to_csv_bytes(_ENUM_HEADERS, values)


def import_enum(name, file_storage, mode):
    headers, records = read_uploaded_table(file_storage)
    parsed = []
    for r in records:
        prop = str(r.get("property", "") or "").strip()
        if not prop:
            continue
        try:
            value = int(float(r.get("value", 0) or 0))
        except (TypeError, ValueError):
            value = 0
        parsed.append({
            "property": prop,
            "value": value,
            "description": str(r.get("description") or ""),
        })

    client = APP.test_client()
    if mode == "append":
        res = client.get(f"/api/enum/{name}")
        existing = res.get_json() if res.status_code == 200 else []
        existing_props = {v.get("property") for v in existing}
        merged = existing + [v for v in parsed if v["property"] not in existing_props]
    else:
        merged = parsed

    save_res = client.post(f"/api/enum/{name}", json=merged)
    ok = 200 <= save_res.status_code < 300
    payload = save_res.get_json(silent=True) or {}
    return ok, len(merged), payload.get("error") or payload.get("message")


# ---------------------------------------------------------------
# ClassDataID
# ---------------------------------------------------------------

def _class_data_id_headers(columns):
    names = ["id", "enum_property"]
    for col in columns:
        col_name = col.get("name")
        if col_name and col_name not in names:
            names.append(col_name)
    return names


def export_class_data_id(name, fmt):
    client = APP.test_client()
    res = client.get(f"/api/class-data-id/{name}")
    if res.status_code != 200:
        return None
    data = res.get_json() or {}
    columns = data.get("columns", [])
    rows = data.get("rows", [])
    headers = _class_data_id_headers(columns)
    if fmt == "xlsx":
        return rows_to_xlsx_bytes(headers, rows, name)
    return rows_to_csv_bytes(headers, rows)


def import_class_data_id(name, file_storage, mode):
    client = APP.test_client()
    res = client.get(f"/api/class-data-id/{name}")
    if res.status_code != 200:
        return False, 0, f"既存データの取得に失敗しました: {name}"
    existing = res.get_json() or {}
    columns = existing.get("columns", [])
    existing_rows = existing.get("rows", [])
    col_types = {c.get("name"): c.get("type") for c in columns if c.get("name")}

    headers, records = read_uploaded_table(file_storage)
    parsed_rows = []
    for r in records:
        row = {}
        for col_name, type_name in col_types.items():
            if col_name in r:
                row[col_name] = coerce_cell(r.get(col_name), type_name)
        enum_property = str(r.get("enum_property") or "").strip()
        if enum_property:
            row["enum_property"] = enum_property
        raw_id = r.get("id")
        try:
            row_id = int(float(raw_id)) if raw_id not in (None, "") else None
        except (TypeError, ValueError):
            row_id = None
        if row_id is not None:
            row["id"] = row_id
        parsed_rows.append(row)

    if mode == "append":
        existing_max_id = max([r.get("id", 0) for r in existing_rows] + [0])
        next_id = existing_max_id + 1
        existing_props = {r.get("enum_property") for r in existing_rows if r.get("enum_property")}
        appended = []
        for row in parsed_rows:
            if row.get("enum_property") and row["enum_property"] in existing_props:
                continue  # 既存と同名(enum_property)の行はスキップ（重複防止）
            row["id"] = next_id
            next_id += 1
            appended.append(row)
        merged_rows = existing_rows + appended
    else:
        # replace: id未指定分は1から連番を振り直す
        next_id = 1
        merged_rows = []
        used_ids = {r["id"] for r in parsed_rows if r.get("id") is not None}
        for row in parsed_rows:
            if row.get("id") is None:
                while next_id in used_ids:
                    next_id += 1
                row["id"] = next_id
                used_ids.add(next_id)
            merged_rows.append(row)

    save_res = client.post(f"/api/class-data-id/{name}", json={"columns": columns, "rows": merged_rows})
    ok = 200 <= save_res.status_code < 300
    payload = save_res.get_json(silent=True) or {}
    return ok, len(merged_rows), payload.get("error") or payload.get("message")


# ---------------------------------------------------------------
# ルート登録
# ---------------------------------------------------------------

_EXPORTERS = {
    "enum": export_enum,
    "class_data_id": export_class_data_id,
}
_IMPORTERS = {
    "enum": import_enum,
    "class_data_id": import_class_data_id,
}
_MIME_TYPES = {
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "csv": "text/csv",
}


def register(app, data_dir):
    from flask import request, send_file, jsonify
    init(app, data_dir)

    @app.route("/api/spreadsheet/capabilities", methods=["GET"])
    def spreadsheet_capabilities():
        return jsonify({"xlsx": _HAS_OPENPYXL, "csv": True})

    @app.route("/api/spreadsheet/<category>/<name>/export", methods=["GET"])
    def spreadsheet_export(category, name):
        exporter = _EXPORTERS.get(category)
        if exporter is None:
            return jsonify({"error": f"未対応のカテゴリです: {category}"}), 400
        fmt = request.args.get("format", "xlsx" if _HAS_OPENPYXL else "csv")
        if fmt not in _MIME_TYPES:
            return jsonify({"error": f"未対応の形式です: {fmt}"}), 400
        try:
            data = exporter(name, fmt)
        except RuntimeError as e:
            return jsonify({"error": str(e)}), 400
        if data is None:
            return jsonify({"error": "見つかりません"}), 404
        return send_file(
            io.BytesIO(data),
            mimetype=_MIME_TYPES[fmt],
            as_attachment=True,
            download_name=f"{name}.{fmt}",
        )

    @app.route("/api/spreadsheet/<category>/<name>/import", methods=["POST"])
    def spreadsheet_import(category, name):
        importer = _IMPORTERS.get(category)
        if importer is None:
            return jsonify({"error": f"未対応のカテゴリです: {category}"}), 400
        file_storage = request.files.get("file")
        if not file_storage:
            return jsonify({"error": "ファイルが指定されていません"}), 400
        mode = request.args.get("mode", "replace")
        if mode not in ("replace", "append"):
            return jsonify({"error": f"未対応のmodeです: {mode}"}), 400
        try:
            ok, count, message = importer(name, file_storage, mode)
        except RuntimeError as e:
            return jsonify({"error": str(e)}), 400
        except Exception as e:
            return jsonify({"error": f"取り込みに失敗しました: {e}"}), 500
        if not ok:
            return jsonify({"error": message or "保存に失敗しました"}), 500
        return jsonify({"message": f"{count}件を取り込みました（{name}）", "count": count})
